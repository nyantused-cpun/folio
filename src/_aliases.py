# -*- coding: utf-8 -*-
"""客户隔离别名：全局 + 客户专属，客户覆盖全局同名 key。
支持：加载/合并/展开文本/对比差异/异议提醒。"""
import os
import json
import yaml

from _paths import SCRIPT_DIR, CLIENTS_DIR

GLOBAL_ALIASES_PATH = os.path.join(SCRIPT_DIR, "_knowledge", "me", "aliases-global.yml")


def _load_yaml(path):
    """加载 yaml，文件不存在返回空 dict。"""
    if not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def load_global():
    """加载全局别名。返回 {标准词: [别名...], ...}。"""
    data = _load_yaml(GLOBAL_ALIASES_PATH)
    return {k: v for k, v in data.items() if isinstance(v, list)}


def load_client(client_name):
    """加载客户专属别名（与全局合并，客户覆盖全局同名 key）。

    支持可维护的词库继承：
    - extends: "全局"  → 继承全局别名（默认行为，已隐式生效）
    - extends: "客户名" → 先加载指定客户的别名作为基础，再覆盖当前客户
    - extends: ["全局", "客户名"] → 多继承，后者覆盖前者
    """
    client_path = os.path.join(CLIENTS_DIR, client_name, "aliases.yml")
    if not os.path.exists(client_path):
        return load_global()

    data = _load_yaml(client_path)
    extends = data.get("extends")

    # 解析 extends：构建继承链
    if extends is None:
        merged = load_global()
    elif isinstance(extends, str):
        merged = _load_base_aliases(extends)
    elif isinstance(extends, list):
        merged = {}
        for src in extends:
            base = _load_base_aliases(src)
            merged.update(base)
    else:
        merged = load_global()

    # 当前客户别名覆盖继承的基础
    # 跳过保留字段：canonical/aliases/markers 是客户身份字段，不是系统同义词
    _RESERVED_KEYS = {"extends", "canonical", "aliases", "markers", "client", "note", "system_names", "vendor_names"}
    for k, v in data.items():
        if k in _RESERVED_KEYS:
            continue
        if isinstance(v, list):
            merged[k] = v
    return merged


def _load_base_aliases(source):
    """加载继承源：'全局' → 全局别名，其他 → 指定客户别名。"""
    if source == "全局":
        return load_global()
    # 视为客户名，递归加载该客户别名
    from _paths import _validate_client_name
    _validate_client_name(source)
    other_path = os.path.join(CLIENTS_DIR, source, "aliases.yml")
    if os.path.exists(other_path):
        return load_client(source)
    # 源不存在则回退全局
    return load_global()


def expand_text(text, aliases):
    """展开文本：对每个 key，若文本含 key 或其任一别名，追加所有同义词。
    aliases: {标准词: [别名...], ...}"""
    if not text or not aliases:
        return text or ""
    additions = []
    for key, alts in aliases.items():
        all_forms = [key] + list(alts)
        if any(form in text for form in all_forms):
            additions.extend(all_forms)
    if additions:
        return text + " " + " ".join(additions)
    return text


def diff_clients(client_a, client_b):
    """对比两客户别名差异。返回 {
        'only_a': {key: alts},
        'only_b': {key: alts},
        'diff_value': {key: {'a': alts, 'b': alts}},
    }"""
    a = load_client(client_a)
    b = load_client(client_b)
    keys_a = set(a.keys())
    keys_b = set(b.keys())
    only_a = {k: a[k] for k in keys_a - keys_b}
    only_b = {k: b[k] for k in keys_b - keys_a}
    diff_value = {}
    for k in keys_a & keys_b:
        if set(a[k]) != set(b[k]):
            diff_value[k] = {"a": a[k], "b": b[k]}
    return {"only_a": only_a, "only_b": only_b, "diff_value": diff_value}


def diff_with_global(client_name):
    """对比客户别名与全局的差异。返回 {
        'overrides': {key: {'global': alts, 'client': alts}},
        'client_only': {key: alts},
    }"""
    global_aliases = load_global()
    client_aliases = load_client(client_name)
    overrides = {}
    client_only = {}
    for k, v in client_aliases.items():
        if k in global_aliases and set(global_aliases[k]) != set(v):
            overrides[k] = {"global": global_aliases[k], "client": v}
        elif k not in global_aliases:
            client_only[k] = v
    return {"overrides": overrides, "client_only": client_only}


def get_client_aliases(client_name):
    """读取客户 aliases.yml 的 aliases 字段（新格式）+ 兼容旧格式 {client: [list]}。
    返回别名列表（不含目录名本身）。
    """
    client_path = os.path.join(CLIENTS_DIR, client_name, "aliases.yml")
    if not os.path.exists(client_path):
        return []
    data = _load_yaml(client_path)
    result = []
    # 新格式：aliases: [list]
    aliases_field = data.get("aliases")
    if isinstance(aliases_field, list):
        for a in aliases_field:
            if a:
                # 兼容 "中心  # 历史旧称" 行内注释（YAML 通常已剥离，保险处理）
                result.append(str(a).split("#")[0].strip())
    # 兼容旧格式：{client_name: [list]}
    old = data.get(client_name)
    if isinstance(old, list):
        result.extend(old)
    # 去重保序，过滤空串
    return [a for a in dict.fromkeys(result) if a]


def resolve_client_name(input_name):
    """解析用户输入的客户名，返回 (resolved, candidates, matched_by)。

    - resolved: 解析后的规范目录名（无匹配时 = input_name）
    - candidates: 近似候选列表（子串匹配但非精确/别名命中）
    - matched_by: 'exact' | 'alias' | 'substring' | None

    用途：session-start / save 前纠错，避免把别名当新客户创建空目录。
    """
    if not input_name:
        return input_name, [], None
    clients = _list_clients()
    # 1. 精确匹配目录名
    if input_name in clients:
        return input_name, [], 'exact'
    # 2. 别名匹配（aliases 字段或旧格式 {client: [list]}）
    for client in clients:
        if input_name in get_client_aliases(client):
            return client, [], 'alias'
    # 3. 子串近似（input 是某目录名子串，或反之）
    candidates = []
    for client in clients:
        if input_name != client and (input_name in client or client in input_name):
            candidates.append(client)
    if candidates:
        return input_name, candidates, 'substring'
    # 4. 无匹配
    return input_name, [], None


def detect_client_in_query(query):
    """从查询文本中检测客户名。返回客户名或 None。"""
    if not os.path.isdir(CLIENTS_DIR):
        return None
    clients = _list_clients()
    for client in clients:
        if client in query:
            return client
        client_aliases = get_client_aliases(client)
        if client_aliases and any(alt in query for alt in client_aliases):
            return client
    return None


def print_alias_reminder(target_client, prev_client=None):
    """打印别名异议提醒。"""
    print("\n⚠️  别名异议提醒:")
    has_reminder = False

    diff_g = diff_with_global(target_client)
    if diff_g["overrides"]:
        has_reminder = True
        print(f"  [{target_client}] 覆盖全局定义:")
        for k, v in diff_g["overrides"].items():
            print(f"    {k}:")
            print(f"      全局: {v['global']}")
            print(f"      {target_client}: {v['client']}")
    if diff_g["client_only"]:
        has_reminder = True
        print(f"  [{target_client}] 独有别名:")
        for k, v in diff_g["client_only"].items():
            print(f"    {k}: {v}")

    if prev_client and prev_client != target_client:
        diff_c = diff_clients(prev_client, target_client)
        if diff_c["diff_value"]:
            has_reminder = True
            print(f"  与上次项目 [{prev_client}] 含义不同:")
            for k, v in diff_c["diff_value"].items():
                print(f"    {k}:")
                print(f"      {prev_client}: {v['a']}")
                print(f"      {target_client}: {v['b']}")

    if not has_reminder:
        print("  无异议（与全局/上次项目一致）")
    print()


def _list_clients():
    """列出所有客户目录名（排除 _ 开头）。"""
    if not os.path.isdir(CLIENTS_DIR):
        return []
    return [d for d in os.listdir(CLIENTS_DIR)
            if os.path.isdir(os.path.join(CLIENTS_DIR, d)) and not d.startswith("_")]


def export_to_excel(path):
    """导出全部别名到 Excel（含表头和现有数据）。

    表头：标准词 | 别名 | 适用范围
    适用范围填"全局"或客户名。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "别名库"

    # 表头
    headers = ["标准词", "别名", "适用范围"]
    ws.append(headers)
    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 全局别名
    global_aliases = load_global()
    for key in sorted(global_aliases.keys()):
        ws.append([key, "、".join(global_aliases[key]), "全局"])

    # 各客户别名
    for client in sorted(_list_clients()):
        load_client(client)
        # 只导出客户独有的（非继承自全局）
        client_path = os.path.join(CLIENTS_DIR, client, "aliases.yml")
        raw = _load_yaml(client_path)
        for key in sorted(raw.keys()):
            if key == "extends":
                continue
            val = raw[key]
            if isinstance(val, list):
                ws.append([key, "、".join(val), client])

    # 自动列宽
    for col in range(1, 4):
        max_len = max(len(str(ws.cell(row=r, column=col).value or ""))
                      for r in range(1, ws.max_row + 1))
        ws.column_dimensions[chr(64 + col)].width = max(max_len + 4, 12)

    wb.save(path)
    total = ws.max_row - 1
    print(f"[export] 导出 {total} 条别名到 {path}")
    return total


def import_from_excel(path):
    """从 Excel 导入别名（合并模式，不删除现有别名）。

    Excel 格式：标准词 | 别名 | 适用范围
    - 标准词：必须，字符串
    - 别名：必须，逗号或顿号分隔
    - 适用范围：必须，"全局"或客户名

    支持 .xlsx 和 .csv（自动检测 UTF-8/GBK 编码）。
    """
    rows = _read_excel_rows(path)

    # 按适用范围分组
    global_new = {}
    client_new = {}  # {client_name: {key: [alts]}}

    for row in rows:
        if not row or not row[0]:
            continue
        key = str(row[0]).strip()
        alts_str = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        scope = str(row[2]).strip() if len(row) > 2 and row[2] else "全局"

        # 别名按逗号或顿号分割
        alts = [a.strip() for a in alts_str.replace(",", "、").split("、") if a.strip()]

        if scope == "全局":
            if key in global_new:
                global_new[key].extend(alts)
                global_new[key] = list(dict.fromkeys(global_new[key]))  # 去重保序
            else:
                global_new[key] = alts
        else:
            if scope not in client_new:
                client_new[scope] = {}
            if key in client_new[scope]:
                client_new[scope][key].extend(alts)
                client_new[scope][key] = list(dict.fromkeys(client_new[scope][key]))
            else:
                client_new[scope][key] = alts

    # 写入全局
    if global_new:
        existing = load_global()
        for key, alts in global_new.items():
            if key in existing:
                merged = list(dict.fromkeys(existing[key] + alts))
                existing[key] = merged
            else:
                existing[key] = alts
        _save_yaml(GLOBAL_ALIASES_PATH, existing)
        print(f"[import] 全局别名: 新增/更新 {len(global_new)} 条")

    # 写入客户
    from _paths import _validate_client_name
    for client_name, entries in client_new.items():
        _validate_client_name(client_name)  # Excel 第三列是外部输入，防路径穿越
        client_path = os.path.join(CLIENTS_DIR, client_name, "aliases.yml")
        existing = _load_yaml(client_path)

        for key, alts in entries.items():
            if key in existing and isinstance(existing[key], list):
                merged = list(dict.fromkeys(existing[key] + alts))
                existing[key] = merged
            else:
                existing[key] = alts

        # 确保客户目录存在
        os.makedirs(os.path.join(CLIENTS_DIR, client_name), exist_ok=True)
        _save_yaml(client_path, existing)
        print(f"[import] {client_name}: 新增/更新 {len(entries)} 条")

    total = len(global_new) + sum(len(v) for v in client_new.values())
    print(f"[import] 完成，共 {total} 条")
    return total


def _save_yaml(path, data):
    """保存 YAML（保留 extends 字段，自动加注释头）。"""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        # 全局文件加注释头
        if path == GLOBAL_ALIASES_PATH:
            f.write("# 全局别名 - 所有客户共享（行业通用术语）\n")
        else:
            client_name = os.path.basename(os.path.dirname(path))
            f.write(f"# {client_name} - 客户别名映射\n")
            if "extends" in data:
                f.write(f"extends: {data['extends']}\n\n")
            else:
                f.write("extends: global\n\n")
        for key in sorted(data.keys()):
            if key == "extends":
                continue
            val = data[key]
            if isinstance(val, list):
                f.write(f"{key}: {val}\n")
            else:
                f.write(f"{key}: {val}\n")


def _read_excel_rows(path):
    """读取 Excel/CSV 文件，返回数据行（跳过表头）。

    支持 .xlsx 和 .csv（自动检测 UTF-8/GBK 编码）。
    """
    ext = os.path.splitext(path)[1].lower()

    if ext == ".csv":
        # CSV：尝试 UTF-8，失败降级 GBK
        for encoding in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
            try:
                with open(path, "r", encoding=encoding) as f:
                    first_line = f.readline()
                    # 验证能正确读取中文
                    if "标准" in first_line or "别名" in first_line or "适用" in first_line or first_line.strip():
                        f.seek(0)
                        lines = f.readlines()
                        # 跳过表头，解析 CSV
                        import csv as _csv
                        reader = _csv.reader(lines[1:])
                        return [tuple(row) for row in reader if row and row[0].strip()]
            except (UnicodeDecodeError, UnicodeError):
                continue
        raise ValueError(f"无法解码 CSV 文件 {path}（尝试了 utf-8/gbk/gb18030）")

    # .xlsx / .xls
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    try:
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    finally:
        wb.close()
    return rows


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("别名工具")
        print("用法: python _aliases.py <show|diff> [参数]")
        sys.exit(0)
    if sys.argv[1] == "show" and len(sys.argv) >= 3:
        aliases = load_client(sys.argv[2])
        for k, v in aliases.items():
            print(f"{k}: {v}")
    elif sys.argv[1] == "diff" and len(sys.argv) >= 4:
        d = diff_clients(sys.argv[2], sys.argv[3])
        print(json.dumps(d, ensure_ascii=False, indent=2))
