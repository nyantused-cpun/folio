# -*- coding: utf-8 -*-
"""报价 Excel 渲染器 — 从 QuoteData 渲染 .xlsx。
数据逻辑在 _quote_data.py，本文件只负责填 Excel 格子。
"""
import shutil
import re
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from _quote_data import (
    load_slots, get_template_source
)

# 列映射（统一 B-H）
COL_INDEX = 2      # B
COL_NAME = 3       # C (merge C:D)
COL_ROLE = 5       # E
COL_PRICE = 6      # F
COL_QTY = 7        # G
COL_AMOUNT = 8     # H

FMT_MONEY = '\\¥#,##0;\\¥\\-#,##0'

# 对齐方式
A_LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
A_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
A_RIGHT = Alignment(horizontal='right', vertical='center', wrap_text=False)


class ExcelRenderer:
    """从 QuoteData 渲染 Excel"""

    def render(self, quote_data, template_name, output_path, style_name="enterprise"):
        # §八 3.5：报价 xlsx 写入（含 shutil.copy2 模板复制）过 output/ 白名单，
        # 入口前置校验，避免模板复制落盘后才失败
        from _renderer import _validate_output_path
        _validate_output_path(output_path)

        from _quote_html import _load_quote_style
        style = _load_quote_style(style_name)
        font_name = style["font"]
        free_color = style["free_color"].lstrip("#")
        self._f_data = Font(name=font_name, size=9)
        self._f_bold = Font(name=font_name, size=9, bold=True)
        self._f_red = Font(name=font_name, size=9, color=free_color)

        src = get_template_source(template_name)
        shutil.copy2(src, output_path)
        wb = openpyxl.load_workbook(output_path)
        slots = load_slots(template_name)

        # 填全局槽位
        self._fill_globals(wb, slots, quote_data.metadata)

        # 判断模板类型
        if "sections" in slots:
            self._render_onepage(wb, quote_data, slots)
        elif "sheets" in slots:
            self._render_multi_sheet(wb, quote_data, slots)

        # 自动排版
        for ws in wb.worksheets:
            self._auto_format_worksheet(ws)

        wb.save(output_path)
        return output_path

    def _get_real_cell(self, ws, row, col):
        """处理合并单元格：返回左上角真实 Cell，如是 MergedCell 则找到左上角。"""
        cell = ws.cell(row=row, column=col)
        if not isinstance(cell, openpyxl.cell.cell.MergedCell):
            return cell
        for merged in ws.merged_cells.ranges:
            if (merged.min_row <= row <= merged.max_row and
                    merged.min_col <= col <= merged.max_col):
                return ws.cell(row=merged.min_row, column=merged.min_col)
        return cell

    def _set_cell(self, cell, value=None, font=None, alignment=None,
                  number_format=None, copy_from=None):
        """统一设置单元格值与样式。"""
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            return cell
        if value is not None:
            cell.value = value
        if font is not None:
            cell.font = font
        if alignment is not None:
            cell.alignment = alignment
        if number_format is not None:
            cell.number_format = number_format
        if copy_from is not None:
            cell.font = copy_from.font
            cell.alignment = copy_from.alignment
            cell.number_format = copy_from.number_format
        return cell

    def _cell_ref_to_row_col(self, cell_ref):
        """把 'C2' / 'AB10' 转成 (row, col)"""
        m = re.match(r"([A-Za-z]+)(\d+)", str(cell_ref or ""))
        if not m:
            return None, None
        return int(m.group(2)), openpyxl.utils.column_index_from_string(m.group(1))

    def _fill_globals(self, wb, slots, metadata):
        for key, cfg in slots.get("global_slots", {}).items():
            cell_ref = cfg.get("cell")
            val = metadata.get(key)
            if not cell_ref or val is None or val == "":
                continue
            sheet = cfg.get("sheet", wb.sheetnames[0])
            if sheet not in wb.sheetnames:
                continue
            row, col = self._cell_ref_to_row_col(cell_ref)
            if row is None:
                continue
            c = self._get_real_cell(wb[sheet], row, col)
            if isinstance(c, openpyxl.cell.cell.MergedCell):
                print(f"[警告] 全局槽位 '{key}' 的单元格 {cell_ref} 在合并区域中无法定位左上角，值未写入")
                continue
            self._set_cell(
                c, value=str(val), font=self._f_data, alignment=A_LEFT_WRAP
            )

    def _render_onepage(self, wb, qd, slots):
        ws = wb[wb.sheetnames[0]]
        sec_by_id = {s.section_id: s for s in qd.sections}

        # 填写各 section（B/C/E/F/G 列），金额列 H 留给公式
        for sec_id, sec_cfg in slots.get("sections", {}).items():
            section = sec_by_id.get(sec_id)
            if not section:
                continue
            self._fill_section(ws, section, sec_cfg)

        # 项目总价区域（写 原价/折扣/折后价）
        self._fill_summary_area(ws, qd, slots)

        # 付款方式区域（写 编号/时间/类型/比例，金额列留给公式或写值）
        self._fill_payment_area(ws, qd, slots)

    def _render_multi_sheet(self, wb, qd, slots):
        sec_by_ref = {s.sheet_ref: s for s in qd.sections}
        for sheet_name, sheet_cfg in slots.get("sheets", {}).items():
            section = sec_by_ref.get(sheet_name)
            if not section or sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            self._fill_section(ws, section, sheet_cfg)
            ref = sheet_cfg.get("summary_ref")
            if ref:
                target = ref.get("sheet", "报价汇总")
                if target in wb.sheetnames:
                    row, col = self._cell_ref_to_row_col(ref["cell"])
                    if row is None:
                        continue
                    c = self._get_real_cell(wb[target], row, col)
                    if not isinstance(c, openpyxl.cell.cell.MergedCell):
                        safe_sheet = target.replace("'", "''")
                        c.value = f"='{safe_sheet}'!H{sheet_cfg['total_row']}"
                        c.font = self._f_data
                        c.alignment = A_RIGHT
                        c.number_format = FMT_MONEY

    def _clear_range(self, ws, start_row, end_row, start_col, end_col):
        """清空指定区域的内容和格式残留。"""
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                cell = self._get_real_cell(ws, r, c)
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                cell.value = None
                cell.font = self._f_data
                cell.number_format = 'General'
                cell.alignment = A_LEFT_WRAP

    def _fill_section(self, ws, section, sec_cfg):
        data_start = sec_cfg["data_start"]
        total_row = sec_cfg["total_row"]
        use_formula = sec_cfg.get("use_formula", False)
        data_end = sec_cfg.get("data_end", total_row - 1)

        # 清空数据区域（B-H），防止模板残留导致重复
        self._clear_range(ws, data_start, data_end, COL_INDEX, COL_AMOUNT)

        # 填数据
        overflow = 0
        for i, item in enumerate(section.items):
            row = data_start + i
            if row > data_end:
                overflow += 1
                continue

            self._set_cell(
                self._get_real_cell(ws, row, COL_INDEX),
                value=item.index, font=self._f_data, alignment=A_CENTER
            )
            self._set_cell(
                self._get_real_cell(ws, row, COL_NAME),
                value=item.name, font=self._f_data, alignment=A_LEFT_WRAP
            )
            self._set_cell(
                self._get_real_cell(ws, row, COL_ROLE),
                value=item.role or "", font=self._f_data, alignment=A_LEFT_WRAP
            )

            if item.unit_price:
                self._set_cell(
                    self._get_real_cell(ws, row, COL_PRICE),
                    value=item.unit_price, font=self._f_data,
                    alignment=A_RIGHT, number_format=FMT_MONEY
                )

            self._set_cell(
                self._get_real_cell(ws, row, COL_QTY),
                value=0 if item.is_gift else item.quantity,
                font=self._f_data, alignment=A_CENTER
            )

            # 金额列：公式模式动态写入，否则写计算值
            if use_formula:
                formula = f"=F{row}*G{row}"
                self._set_cell(
                    self._get_real_cell(ws, row, COL_AMOUNT),
                    value=formula,
                    font=self._f_red if item.is_gift else self._f_data,
                    alignment=A_RIGHT, number_format=FMT_MONEY
                )
            else:
                self._set_cell(
                    self._get_real_cell(ws, row, COL_AMOUNT),
                    value=0 if item.is_gift else item.amount,
                    font=self._f_red if item.is_gift else self._f_data,
                    alignment=A_RIGHT, number_format=FMT_MONEY
                )

        # 清空未使用的金额列（避免模板公式残留）
        for r in range(data_start, data_end + 1):
            if r - data_start >= len(section.items):
                self._set_cell(
                    self._get_real_cell(ws, r, COL_AMOUNT),
                    value=None, font=self._f_data, alignment=A_RIGHT
                )

        # 合计行
        if use_formula:
            formula = f"=SUM(H{data_start}:H{data_end})"
            self._set_cell(
                self._get_real_cell(ws, total_row, COL_AMOUNT),
                value=formula, font=self._f_bold,
                alignment=A_RIGHT, number_format=FMT_MONEY
            )
        else:
            total = sum(item.amount for item in section.items)
            self._set_cell(
                self._get_real_cell(ws, total_row, COL_AMOUNT),
                value=total, font=self._f_bold,
                alignment=A_RIGHT, number_format=FMT_MONEY
            )

        if overflow > 0:
            capacity = data_end - data_start + 1
            print(f"[警告] 区块 '{section.title}' 数据行溢出：{len(section.items)} 项，模板预留 {capacity} 行，丢弃 {overflow} 项")

    def _fill_summary_area(self, ws, qd, slots):
        """填充项目总价区域（原价/折扣/折后价）"""
        summary_cfg = slots.get("summary", {})
        row_map = summary_cfg.get("row_map", {})
        total_row = summary_cfg.get("total_row", 8)

        grand_total = 0
        for sec_id, row in row_map.items():
            section = next((s for s in qd.sections if s.section_id == sec_id), None)
            if not section or not section.items:
                continue
            # 原价 (E列)
            self._set_cell(
                self._get_real_cell(ws, row, 5),
                value=section.original_amount, font=self._f_data,
                alignment=A_RIGHT, number_format=FMT_MONEY
            )
            # 折扣文本 (F列)
            self._set_cell(
                self._get_real_cell(ws, row, 6),
                value=section.discount_summary, font=self._f_data,
                alignment=A_CENTER
            )
            # 折后价 (G:H 合并列)
            self._set_cell(
                self._get_real_cell(ws, row, 7),
                value=section.total_amount, font=self._f_data,
                alignment=A_RIGHT, number_format=FMT_MONEY
            )
            grand_total += section.total_amount

        # 合计行 (H列)
        self._set_cell(
            self._get_real_cell(ws, total_row, 8),
            value=grand_total, font=self._f_bold,
            alignment=A_RIGHT, number_format=FMT_MONEY
        )

    def _fill_payment_area(self, ws, qd, slots):
        """填充付款方式区域"""
        payment_cfg = slots.get("payment", {})
        data_start = payment_cfg.get("data_start", 20)
        max_rows = payment_cfg.get("max_rows", 3)

        # 清空所有预留行（B-H）
        self._clear_range(ws, data_start, data_start + max_rows - 1, 2, 8)

        for i, pay in enumerate(qd.payments):
            if i >= max_rows:
                print(f"[警告] 付款方式溢出：{len(qd.payments)} 条，模板预留 {max_rows} 条")
                break
            row = data_start + i
            self._set_cell(
                self._get_real_cell(ws, row, 2),
                value=i + 1, font=self._f_data, alignment=A_CENTER
            )
            self._set_cell(
                self._get_real_cell(ws, row, 3),
                value=pay["time"], font=self._f_data, alignment=A_LEFT_WRAP
            )
            self._set_cell(
                self._get_real_cell(ws, row, 5),
                value=pay["type"], font=self._f_data, alignment=A_LEFT_WRAP
            )
            self._set_cell(
                self._get_real_cell(ws, row, 6),
                value=pay["ratio"], font=self._f_data, alignment=A_CENTER
            )
            self._set_cell(
                self._get_real_cell(ws, row, 7),
                value=pay["amount"], font=self._f_data,
                alignment=A_RIGHT, number_format=FMT_MONEY
            )

    def _auto_format_worksheet(self, ws):
        """自动调整列宽与行高。"""
        # 列宽：根据内容长度，限制在 8~50 之间
        for col_cells in ws.columns:
            cell0 = col_cells[0]
            col_letter = get_column_letter(cell0.column)
            max_len = 0
            for cell in col_cells:
                if cell.value is None:
                    continue
                try:
                    length = len(str(cell.value))
                except Exception:
                    length = 0
                if length > max_len:
                    max_len = length
            adjusted = min(max(max_len + 2, 8), 50)
            ws.column_dimensions[col_letter].width = adjusted

        # 行高：根据换行文本粗略估算
        for row_cells in ws.iter_rows():
            row_num = row_cells[0].row
            max_height = 15
            for cell in row_cells:
                if cell.value is None:
                    continue
                alignment = cell.alignment
                if alignment and alignment.wrap_text:
                    text = str(cell.value)
                    col_letter = get_column_letter(cell.column)
                    col_dim = ws.column_dimensions.get(col_letter)
                    width = col_dim.width if col_dim and getattr(col_dim, "width", None) else 8
                    lines = max(1, int(len(text) / max(width, 1)) + 1)
                    max_height = max(max_height, min(lines * 15, 120))
            ws.row_dimensions[row_num].height = max_height

    # 校验
    def validate(self, xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
        issues = []
        func_re = re.compile(r"^[A-Z]{2,}$")
        sheet_ref_re = re.compile(r"'([^']+)'!|([A-Za-z_][A-Za-z0-9_]*)!")
        for sn in wb.sheetnames:
            for row in wb[sn].iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        for m in sheet_ref_re.finditer(cell.value):
                            ref = m.group(1) or m.group(2)
                            if ref and ref not in wb.sheetnames and not func_re.match(ref):
                                issues.append(f"{sn}!{cell.coordinate}: 引用不存在的sheet '{ref}'")
        return issues
